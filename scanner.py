import os
import json
import re
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone

from pybit.unified_trading import HTTP
import pandas as pd
from tabulate import tabulate
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill


session = HTTP(testnet=False)

MIN_TURNOVER = 50_000_000
MAX_SYMBOLS_TO_SCAN = 30
REQUEST_SLEEP = 0.45
TOP_N = 10
TOP_N_PER_SIDE = 10

ENABLE_LONG_SIGNALS = True
ENABLE_SHORT_SIGNALS = True
OI_LOOKBACK_BARS = 24
OI_CONFIRM_THRESHOLD_PCT = 3.0
PRICE_MOVE_THRESHOLD_PCT = 1.0

ENABLE_SMC_FILTER = True
SMC_INTERNAL_LENGTH = 5
SMC_SWING_LENGTH = 20
SMC_RECENT_EVENT_BARS = 12
SMC_EQUAL_LEVEL_THRESHOLD_ATR = 0.1

ENABLE_MARKET_REGIME_CONTEXT = True
MARKET_REGIME_MANUAL_FILE = "market_regime_manual.json"
HTTP_TIMEOUT = 15

CSV_FILE = "scanner_results.csv"
XLSX_FILE = "scanner_results.xlsx"
SIGNALS_LOG = "signals_log.csv"
SELECTED_SIGNALS_LOG = "selected_signals.csv"
MARKET_REGIME_LOG = "market_regime_log.csv"


def is_yes(value):
    return str(value).strip().lower() in {"yes", "y", "1", "true", "да", "д"}


def http_get_json(url, params=None):
    if params:
        url = url + "?" + urllib.parse.urlencode(params)

    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json,text/plain,*/*",
        },
    )

    with urllib.request.urlopen(request, timeout=HTTP_TIMEOUT) as response:
        return json.loads(response.read().decode("utf-8"))


def http_get_text(url):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )

    with urllib.request.urlopen(request, timeout=HTTP_TIMEOUT) as response:
        return response.read().decode("utf-8", errors="ignore")


def read_manual_market_regime():
    if not os.path.exists(MARKET_REGIME_MANUAL_FILE):
        return {}

    try:
        with open(MARKET_REGIME_MANUAL_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        if not isinstance(data, dict):
            return {}

        return data
    except Exception:
        return {}


def to_float(value):
    if value in {"", None}:
        return None

    try:
        cleaned = str(value)
        cleaned = cleaned.replace("%", "")
        cleaned = cleaned.replace(",", ".")
        cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
        if cleaned in {"", "-", "."}:
            return None
        return float(cleaned)
    except Exception:
        return None


def extract_cmc_cycle_value(html, label):
    # CoinMarketCap pages are dynamic, so this is a best-effort fallback only.
    # If it fails, manual file / other APIs still keep the scanner running.
    pattern = re.compile(
        re.escape(label) + r".{0,800}?([-+]?\d+(?:[.,]\d+)?%?)",
        re.IGNORECASE | re.DOTALL,
    )
    match = pattern.search(html)

    if not match:
        return None

    return to_float(match.group(1))


def classify_fear_greed(value):
    value = to_float(value)
    if value is None:
        return ""
    if value <= 24:
        return "extreme fear"
    if value <= 44:
        return "fear"
    if value <= 55:
        return "neutral"
    if value <= 74:
        return "greed"
    return "extreme greed"


def classify_altseason(value):
    value = to_float(value)
    if value is None:
        return ""
    if value >= 75:
        return "altseason"
    if value <= 25:
        return "bitcoin season"
    return "neutral"


def get_market_regime_context():
    context = {
        "market_regime": "unknown",
        "market_bias": "",
        "market_summary": "",
        "fear_greed_value": "",
        "fear_greed_status": "",
        "altseason_index": "",
        "altseason_status": "",
        "btc_dominance": "",
        "eth_dominance": "",
        "cycle_mvrv_z": "",
        "cycle_puell": "",
        "cycle_mayer": "",
        "cycle_cbbi": "",
        "market_reason": "",
    }

    if not ENABLE_MARKET_REGIME_CONTEXT:
        context["market_regime"] = "disabled"
        return context

    reasons = []

    manual = read_manual_market_regime()

    for key in context:
        if key in manual and manual[key] not in {"", None}:
            context[key] = manual[key]

    try:
        fng = http_get_json("https://api.alternative.me/fng/", {"limit": 1, "format": "json"})
        row = fng.get("data", [{}])[0]
        if row:
            context["fear_greed_value"] = to_float(row.get("value"))
            context["fear_greed_status"] = str(row.get("value_classification", "")) or classify_fear_greed(context["fear_greed_value"])
    except Exception as e:
        reasons.append(f"Fear/Greed unavailable: {e}")

    try:
        global_data = http_get_json("https://api.coingecko.com/api/v3/global")
        percentages = global_data.get("data", {}).get("market_cap_percentage", {})
        if "btc" in percentages:
            context["btc_dominance"] = round(float(percentages["btc"]), 2)
        if "eth" in percentages:
            context["eth_dominance"] = round(float(percentages["eth"]), 2)
    except Exception as e:
        reasons.append(f"dominance unavailable: {e}")

    try:
        html = http_get_text("https://coinmarketcap.com/ru/charts/crypto-market-cycle-indicators/")

        cmc_altseason = extract_cmc_cycle_value(html, "CMC Altcoin Season Index")
        cmc_btc_dom = extract_cmc_cycle_value(html, "Bitcoin Dominance")
        cmc_mvrv_z = extract_cmc_cycle_value(html, "MVRV Z-Score")
        cmc_puell = extract_cmc_cycle_value(html, "Puell Multiple")
        cmc_mayer = extract_cmc_cycle_value(html, "Mayer Multiple")
        cmc_cbbi = extract_cmc_cycle_value(html, "Crypto Bitcoin Bull Run Index")

        if cmc_altseason is not None and context["altseason_index"] == "":
            context["altseason_index"] = cmc_altseason
        if cmc_btc_dom is not None and context["btc_dominance"] == "":
            context["btc_dominance"] = cmc_btc_dom
        if cmc_mvrv_z is not None:
            context["cycle_mvrv_z"] = cmc_mvrv_z
        if cmc_puell is not None:
            context["cycle_puell"] = cmc_puell
        if cmc_mayer is not None:
            context["cycle_mayer"] = cmc_mayer
        if cmc_cbbi is not None:
            context["cycle_cbbi"] = cmc_cbbi
    except Exception as e:
        reasons.append(f"CMC cycle unavailable: {e}")

    if context["fear_greed_status"] == "":
        context["fear_greed_status"] = classify_fear_greed(context["fear_greed_value"])

    if context["altseason_status"] == "":
        context["altseason_status"] = classify_altseason(context["altseason_index"])

    fng_value = to_float(context["fear_greed_value"])
    alt_value = to_float(context["altseason_index"])
    btc_dom_value = to_float(context["btc_dominance"])
    mvrv_value = to_float(context["cycle_mvrv_z"])
    puell_value = to_float(context["cycle_puell"])

    if context["altseason_status"]:
        reasons.append(f"altseason: {context['altseason_status']}")

    if context["fear_greed_status"]:
        reasons.append(f"sentiment: {context['fear_greed_status']}")

    if btc_dom_value is not None:
        if btc_dom_value >= 58:
            reasons.append("BTC dominance high: alt longs need extra confirmation")
        elif btc_dom_value <= 50:
            reasons.append("BTC dominance lower: alt risk appetite better")

    if fng_value is not None:
        if fng_value >= 75:
            reasons.append("extreme greed: avoid chasing late longs")
        elif fng_value <= 25:
            reasons.append("extreme fear: avoid late shorts into support")

    if alt_value is not None:
        if alt_value >= 75:
            reasons.append("altseason supports selective alt longs")
        elif alt_value <= 25:
            reasons.append("bitcoin season: alt longs lower quality")

    if mvrv_value is not None and mvrv_value >= 5:
        reasons.append("cycle MVRV elevated")

    if puell_value is not None and puell_value >= 3:
        reasons.append("Puell elevated")

    if context["market_regime"] == "unknown":
        if fng_value is not None and fng_value >= 75:
            context["market_regime"] = "risk-on overheated"
        elif fng_value is not None and fng_value <= 25:
            context["market_regime"] = "risk-off fearful"
        elif alt_value is not None and alt_value >= 75:
            context["market_regime"] = "altseason"
        elif alt_value is not None and alt_value <= 25:
            context["market_regime"] = "bitcoin season"
        else:
            context["market_regime"] = "neutral/mixed"

    context["market_reason"] = "; ".join(reasons)
    context.update(build_market_summary(context))

    return context


def build_market_summary(context):
    fng_value = to_float(context.get("fear_greed_value", ""))
    alt_value = to_float(context.get("altseason_index", ""))
    btc_dom_value = to_float(context.get("btc_dominance", ""))
    mvrv_value = to_float(context.get("cycle_mvrv_z", ""))
    puell_value = to_float(context.get("cycle_puell", ""))
    cbbi_value = to_float(context.get("cycle_cbbi", ""))

    bearish_points = 0
    bullish_points = 0
    caution_points = 0
    notes = []

    if fng_value is not None:
        if fng_value <= 25:
            bearish_points += 1
            caution_points += 1
            notes.append("страх на рынке: поздние шорты у поддержки опасны")
        elif fng_value >= 75:
            bullish_points += 1
            caution_points += 1
            notes.append("жадность на рынке: поздние лонги уязвимы к откату")
        elif fng_value >= 55:
            bullish_points += 1
            notes.append("умеренный risk-on по сентименту")
        elif fng_value <= 44:
            bearish_points += 1
            notes.append("умеренный risk-off по сентименту")

    if alt_value is not None:
        if alt_value >= 75:
            bullish_points += 2
            notes.append("альтсезон поддерживает выборочные long по альтам")
        elif alt_value <= 25:
            bearish_points += 1
            notes.append("bitcoin season снижает качество alt-long")

    if btc_dom_value is not None:
        if btc_dom_value >= 58:
            bearish_points += 1
            notes.append("доминирование BTC высокое: alt-long требуют усиленного подтверждения")
        elif btc_dom_value <= 50:
            bullish_points += 1
            notes.append("доминирование BTC ниже: фон лучше для alt-risk")

    if mvrv_value is not None and mvrv_value >= 5:
        caution_points += 2
        notes.append("MVRV повышен: есть риск перегрева цикла")

    if puell_value is not None and puell_value >= 3:
        caution_points += 2
        notes.append("Puell повышен: late-cycle long опаснее")

    if cbbi_value is not None:
        if cbbi_value >= 80:
            caution_points += 2
            bullish_points += 1
            notes.append("CBBI высокий: потенциал bull-cycle уже менее выгоден по R/R")
        elif cbbi_value <= 35:
            bearish_points += 1
            notes.append("CBBI низкий/средний: эйфории рынка нет")

    if bullish_points - bearish_points >= 2 and caution_points <= 1:
        market_bias = "risk-on / bullish"
        summary = "Рынок выглядит скорее risk-on: long-сигналы можно рассматривать активнее, но только от уровней и без погони."
    elif bearish_points - bullish_points >= 2:
        market_bias = "risk-off / bearish"
        summary = "Рынок выглядит скорее risk-off: short-сигналы приоритетнее, а long по альтам требуют сильного подтверждения и меньшего размера."
    elif caution_points >= 2 and bullish_points >= bearish_points:
        market_bias = "risk-on but overheated"
        summary = "Рынок не обязательно медвежий, но есть признаки перегрева: long лучше брать только после отката, late-entry стоит избегать."
    elif caution_points >= 2 and bearish_points > bullish_points:
        market_bias = "bearish with reversal risk"
        summary = "Фон слабый, но есть риск резких отскоков: шорты у поддержки опасны, лучше ждать ретесты сопротивлений."
    else:
        market_bias = "mixed / neutral"
        summary = "Рыночный фон смешанный: нет явного преимущества для long или short, решают BTC-фильтр, локальная структура и качество входа."

    if notes:
        summary = summary + " Факторы: " + "; ".join(notes[:4]) + "."

    return {
        "market_bias": market_bias,
        "market_summary": summary,
    }


def print_market_regime_context(context):
    print("Market context:")
    print(f"  regime: {context.get('market_regime', '')}")
    print(f"  bias: {context.get('market_bias', '')}")
    print(f"  fear/greed: {context.get('fear_greed_value', '')} {context.get('fear_greed_status', '')}")
    print(f"  altseason: {context.get('altseason_index', '')} {context.get('altseason_status', '')}")
    print(f"  BTC dominance: {context.get('btc_dominance', '')}")
    print(f"  ETH dominance: {context.get('eth_dominance', '')}")

    cycle_parts = []

    if context.get("cycle_mvrv_z", "") != "":
        cycle_parts.append(f"MVRV Z={context.get('cycle_mvrv_z')}")

    if context.get("cycle_puell", "") != "":
        cycle_parts.append(f"Puell={context.get('cycle_puell')}")

    if context.get("cycle_mayer", "") != "":
        cycle_parts.append(f"Mayer={context.get('cycle_mayer')}")

    if context.get("cycle_cbbi", "") != "":
        cycle_parts.append(f"CBBI={context.get('cycle_cbbi')}")

    if cycle_parts:
        print(f"  cycle: {', '.join(cycle_parts)}")

    if context.get("market_reason"):
        print(f"  summary: {context.get('market_reason')}")

    if context.get("market_summary"):
        print("")
        print(f"Market summary: {context.get('market_summary')}")

    print("")


def append_market_regime_log(context):
    if not ENABLE_MARKET_REGIME_CONTEXT:
        return

    row = context.copy()
    row["created_at"] = datetime.now(timezone.utc).isoformat()

    new_df = pd.DataFrame([row])

    if os.path.exists(MARKET_REGIME_LOG):
        old_df = pd.read_csv(MARKET_REGIME_LOG)
        final_df = pd.concat([old_df, new_df], ignore_index=True)
    else:
        final_df = new_df

    final_df.to_csv(MARKET_REGIME_LOG, index=False, encoding="utf-8-sig")


def ema(series, period):
    return series.ewm(span=period, adjust=False).mean()


def rsi(series, period=14):
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()

    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def atr(df, period=14):
    prev_close = df["close"].shift(1)

    tr = pd.concat([
        df["high"] - df["low"],
        (df["high"] - prev_close).abs(),
        (df["low"] - prev_close).abs(),
    ], axis=1).max(axis=1)

    return tr.rolling(period).mean()


def find_confirmed_pivots(df, size):
    pivots = []

    if len(df) < size * 2 + 3:
        return pivots

    highs = df["high"].tolist()
    lows = df["low"].tolist()

    # Confirmed pivots only: do not use future bars beyond the confirmation window.
    for i in range(size, len(df) - size):
        high_window = highs[i - size:i + size + 1]
        low_window = lows[i - size:i + size + 1]

        if highs[i] == max(high_window):
            pivots.append({
                "index": i,
                "type": "high",
                "level": highs[i],
                "timestamp": df.iloc[i]["timestamp"],
            })

        if lows[i] == min(low_window):
            pivots.append({
                "index": i,
                "type": "low",
                "level": lows[i],
                "timestamp": df.iloc[i]["timestamp"],
            })

    pivots.sort(key=lambda x: x["index"])
    return pivots


def detect_structure_events(df, size, structure_name):
    pivots = find_confirmed_pivots(df, size)
    events = []
    trend_bias = 0

    last_high = None
    last_low = None
    high_crossed = True
    low_crossed = True

    pivots_by_index = {}
    for pivot in pivots:
        pivots_by_index.setdefault(pivot["index"], []).append(pivot)

    for i, row in df.iterrows():
        for pivot in pivots_by_index.get(i, []):
            if pivot["type"] == "high":
                last_high = pivot
                high_crossed = False
            elif pivot["type"] == "low":
                last_low = pivot
                low_crossed = False

        close = float(row["close"])

        if last_high and not high_crossed and close > float(last_high["level"]):
            tag = "CHoCH" if trend_bias == -1 else "BOS"
            trend_bias = 1
            high_crossed = True
            events.append({
                "index": i,
                "structure": structure_name,
                "direction": "bullish",
                "tag": tag,
                "level": float(last_high["level"]),
            })

        if last_low and not low_crossed and close < float(last_low["level"]):
            tag = "CHoCH" if trend_bias == 1 else "BOS"
            trend_bias = -1
            low_crossed = True
            events.append({
                "index": i,
                "structure": structure_name,
                "direction": "bearish",
                "tag": tag,
                "level": float(last_low["level"]),
            })

    return events, pivots, trend_bias


def detect_recent_fvg(df, lookback=30):
    events = []
    start = max(2, len(df) - lookback)

    for i in range(start, len(df)):
        current_low = float(df.iloc[i]["low"])
        current_high = float(df.iloc[i]["high"])
        prev_close = float(df.iloc[i - 1]["close"])
        high_2 = float(df.iloc[i - 2]["high"])
        low_2 = float(df.iloc[i - 2]["low"])

        if current_low > high_2 and prev_close > high_2:
            events.append({
                "index": i,
                "direction": "bullish",
                "top": current_low,
                "bottom": high_2,
            })

        if current_high < low_2 and prev_close < low_2:
            events.append({
                "index": i,
                "direction": "bearish",
                "top": low_2,
                "bottom": current_high,
            })

    return events


def analyze_smc(df4h, side):
    if not ENABLE_SMC_FILTER:
        return {
            "smc_score_delta": 0,
            "smc_bias": "disabled",
            "smc_event": "",
            "smc_zone": "",
            "smc_reason": "",
        }

    try:
        internal_events, internal_pivots, internal_bias = detect_structure_events(
            df4h,
            SMC_INTERNAL_LENGTH,
            "internal",
        )
        swing_events, swing_pivots, swing_bias = detect_structure_events(
            df4h,
            SMC_SWING_LENGTH,
            "swing",
        )

        all_events = internal_events + swing_events
        all_events.sort(key=lambda x: x["index"])
        recent_events = [
            event for event in all_events
            if len(df4h) - 1 - int(event["index"]) <= SMC_RECENT_EVENT_BARS
        ]

        latest_event = recent_events[-1] if recent_events else (all_events[-1] if all_events else None)

        swing_highs = [p for p in swing_pivots if p["type"] == "high"]
        swing_lows = [p for p in swing_pivots if p["type"] == "low"]

        last_price = float(df4h.iloc[-1]["close"])
        last_swing_high = swing_highs[-1]["level"] if swing_highs else df4h.tail(60)["high"].max()
        last_swing_low = swing_lows[-1]["level"] if swing_lows else df4h.tail(60)["low"].min()

        range_size = float(last_swing_high) - float(last_swing_low)
        zone = "unknown"

        if range_size > 0:
            position_in_range = (last_price - float(last_swing_low)) / range_size

            if position_in_range >= 0.70:
                zone = "premium"
            elif position_in_range <= 0.30:
                zone = "discount"
            else:
                zone = "equilibrium"

        fvg_events = detect_recent_fvg(df4h)
        latest_fvg = fvg_events[-1] if fvg_events else None

        score_delta = 0
        reasons = []

        if side == "long":
            if latest_event and latest_event["direction"] == "bullish":
                boost = 2 if latest_event["tag"] == "CHoCH" else 1
                score_delta += boost
                reasons.append(f"SMC bullish {latest_event['structure']} {latest_event['tag']}")

            if latest_event and latest_event["direction"] == "bearish":
                score_delta -= 2
                reasons.append(f"SMC bearish {latest_event['structure']} {latest_event['tag']}")

            if zone == "discount":
                score_delta += 1
                reasons.append("SMC discount zone")
            elif zone == "premium":
                score_delta -= 1
                reasons.append("SMC premium zone")

            if latest_fvg and latest_fvg["direction"] == "bullish":
                score_delta += 1
                reasons.append("SMC bullish FVG")
            elif latest_fvg and latest_fvg["direction"] == "bearish":
                score_delta -= 1
                reasons.append("SMC bearish FVG")

        elif side == "short":
            if latest_event and latest_event["direction"] == "bearish":
                boost = 2 if latest_event["tag"] == "CHoCH" else 1
                score_delta += boost
                reasons.append(f"SMC bearish {latest_event['structure']} {latest_event['tag']}")

            if latest_event and latest_event["direction"] == "bullish":
                score_delta -= 2
                reasons.append(f"SMC bullish {latest_event['structure']} {latest_event['tag']}")

            if zone == "premium":
                score_delta += 1
                reasons.append("SMC premium zone")
            elif zone == "discount":
                score_delta -= 1
                reasons.append("SMC discount zone")

            if latest_fvg and latest_fvg["direction"] == "bearish":
                score_delta += 1
                reasons.append("SMC bearish FVG")
            elif latest_fvg and latest_fvg["direction"] == "bullish":
                score_delta -= 1
                reasons.append("SMC bullish FVG")

        if swing_bias == 1:
            smc_bias = "swing bullish"
        elif swing_bias == -1:
            smc_bias = "swing bearish"
        elif internal_bias == 1:
            smc_bias = "internal bullish"
        elif internal_bias == -1:
            smc_bias = "internal bearish"
        else:
            smc_bias = "neutral"

        smc_event = ""
        if latest_event:
            bars_ago = len(df4h) - 1 - int(latest_event["index"])
            smc_event = (
                f"{latest_event['structure']} "
                f"{latest_event['direction']} "
                f"{latest_event['tag']} "
                f"{bars_ago} bars ago"
            )

        return {
            "smc_score_delta": score_delta,
            "smc_bias": smc_bias,
            "smc_event": smc_event,
            "smc_zone": zone,
            "smc_reason": ", ".join(reasons),
        }

    except Exception as e:
        return {
            "smc_score_delta": 0,
            "smc_bias": "unavailable",
            "smc_event": "",
            "smc_zone": "",
            "smc_reason": f"SMC unavailable: {e}",
        }


def get_ticker_map():
    data = session.get_tickers(category="linear")
    return {x["symbol"]: x for x in data["result"]["list"]}


def get_klines(symbol, interval="240", limit=200):
    data = session.get_kline(
        category="linear",
        symbol=symbol,
        interval=interval,
        limit=limit,
    )

    rows = data["result"]["list"]
    rows = list(reversed(rows))

    df = pd.DataFrame(rows, columns=[
        "timestamp",
        "open",
        "high",
        "low",
        "close",
        "volume",
        "turnover",
    ])

    for col in ["open", "high", "low", "close", "volume", "turnover"]:
        df[col] = df[col].astype(float)

    df["timestamp"] = pd.to_datetime(df["timestamp"].astype(int), unit="ms")

    return df


def get_open_interest(symbol, interval_time="4h", limit=24):
    data = session.get_open_interest(
        category="linear",
        symbol=symbol,
        intervalTime=interval_time,
        limit=limit,
    )

    rows = data.get("result", {}).get("list", [])

    if len(rows) < 2:
        return None

    df = pd.DataFrame(rows)
    df["openInterest"] = df["openInterest"].astype(float)
    df["timestamp"] = df["timestamp"].astype(int)
    df = df.sort_values("timestamp")

    return df


def get_btc_context():
    try:
        df = get_klines("BTCUSDT", "240", 200)

        df["ema20"] = ema(df["close"], 20)
        df["ema50"] = ema(df["close"], 50)
        df["atr"] = atr(df, 14)

        last = df.iloc[-1]

        previous_structure = df.iloc[-35:-3]
        previous_swing_low = previous_structure["low"].min()
        previous_swing_high = previous_structure["high"].max()

        btc_price = float(last["close"])
        btc_ema20 = float(last["ema20"])
        btc_ema50 = float(last["ema50"])

        hard_break = (
            btc_price < btc_ema50
            and float(last["low"]) < previous_swing_low
        )

        bearish_trend = (
            btc_price < btc_ema50
            and btc_ema20 < btc_ema50
        )

        weak_context = (
            btc_price < btc_ema20
            or btc_price < btc_ema50
        )

        if hard_break:
            regime = "BTC 4H structure broken"
            allow_longs = True
            allow_shorts = True
            long_score_delta = -3
            short_score_delta = 2
        elif bearish_trend:
            regime = "BTC 4H bearish trend"
            allow_longs = True
            allow_shorts = True
            long_score_delta = -2
            short_score_delta = 2
        elif weak_context:
            regime = "BTC 4H weak/neutral"
            allow_longs = True
            allow_shorts = True
            long_score_delta = -1
            short_score_delta = 1
        else:
            regime = "BTC 4H supportive"
            allow_longs = True
            allow_shorts = True
            long_score_delta = 1
            short_score_delta = -2

        return {
            "btc_price": round(btc_price, 2),
            "btc_ema20": round(btc_ema20, 2),
            "btc_ema50": round(btc_ema50, 2),
            "previous_swing_low": round(previous_swing_low, 2),
            "previous_swing_high": round(previous_swing_high, 2),
            "regime": regime,
            "allow_longs": allow_longs,
            "allow_shorts": allow_shorts,
            "long_score_delta": long_score_delta,
            "short_score_delta": short_score_delta,
        }

    except Exception as e:
        return {
            "btc_price": "",
            "btc_ema20": "",
            "btc_ema50": "",
            "previous_swing_low": "",
            "previous_swing_high": "",
            "regime": f"BTC filter unavailable: {e}",
            "allow_longs": True,
            "allow_shorts": True,
            "long_score_delta": 0,
            "short_score_delta": 0,
        }


def analyze_open_interest(symbol, df4h, side):
    try:
        oi_df = get_open_interest(
            symbol=symbol,
            interval_time="4h",
            limit=OI_LOOKBACK_BARS,
        )

        time.sleep(REQUEST_SLEEP)

        if oi_df is None or len(oi_df) < 2:
            return {
                "oi_change_pct": "",
                "price_change_oi_window_pct": "",
                "oi_score_delta": 0,
                "oi_signal": "OI unavailable",
            }

        oi_start = float(oi_df.iloc[0]["openInterest"])
        oi_end = float(oi_df.iloc[-1]["openInterest"])

        if oi_start <= 0:
            return {
                "oi_change_pct": "",
                "price_change_oi_window_pct": "",
                "oi_score_delta": 0,
                "oi_signal": "OI invalid",
            }

        oi_change_pct = (oi_end - oi_start) / oi_start * 100

        price_window = df4h.tail(len(oi_df))

        if len(price_window) < 2:
            price_change_pct = 0
        else:
            price_start = float(price_window.iloc[0]["close"])
            price_end = float(price_window.iloc[-1]["close"])
            price_change_pct = (price_end - price_start) / price_start * 100

        score_delta = 0
        signal = "OI neutral"

        if side == "long":
            if (
                price_change_pct >= PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct >= OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = 1
                signal = "price up + OI up: long confirmation"

            elif (
                price_change_pct <= -PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct >= OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = -2
                signal = "price down + OI up: long pressure"

            elif (
                price_change_pct >= PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct <= -OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = 0
                signal = "price up + OI down: possible short squeeze"

            elif (
                price_change_pct <= -PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct <= -OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = 0
                signal = "price down + OI down: deleveraging"

        elif side == "short":
            if (
                price_change_pct <= -PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct >= OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = 1
                signal = "price down + OI up: short confirmation"

            elif (
                price_change_pct >= PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct >= OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = -2
                signal = "price up + OI up: short pressure"

            elif (
                price_change_pct <= -PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct <= -OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = 0
                signal = "price down + OI down: long liquidation/deleveraging"

            elif (
                price_change_pct >= PRICE_MOVE_THRESHOLD_PCT
                and oi_change_pct <= -OI_CONFIRM_THRESHOLD_PCT
            ):
                score_delta = 0
                signal = "price up + OI down: short covering"

        return {
            "oi_change_pct": round(oi_change_pct, 2),
            "price_change_oi_window_pct": round(price_change_pct, 2),
            "oi_score_delta": score_delta,
            "oi_signal": signal,
        }

    except Exception as e:
        return {
            "oi_change_pct": "",
            "price_change_oi_window_pct": "",
            "oi_score_delta": 0,
            "oi_signal": f"OI unavailable: {e}",
        }


def analyze_symbol(symbol, ticker, btc_context, market_context, side):
    df4h = get_klines(symbol, "240", 200)
    time.sleep(REQUEST_SLEEP)

    df1d = get_klines(symbol, "D", 120)
    time.sleep(REQUEST_SLEEP)

    for df in [df4h, df1d]:
        df["ema20"] = ema(df["close"], 20)
        df["ema50"] = ema(df["close"], 50)
        df["rsi"] = rsi(df["close"], 14)
        df["atr"] = atr(df, 14)

    last = df4h.iloc[-1]
    dlast = df1d.iloc[-1]

    price = float(ticker["lastPrice"])
    turnover24h = float(ticker["turnover24h"])
    change24h = float(ticker["price24hPcnt"]) * 100
    funding = float(ticker.get("fundingRate", 0) or 0) * 100

    if turnover24h < MIN_TURNOVER:
        return None

    recent = df4h.tail(30)

    support = recent["low"].min()
    resistance = recent["high"].max()

    if side == "long":
        stop = min(support * 0.995, price - 1.2 * last["atr"])
        risk = price - stop

        if risk <= 0:
            return None

        target_by_level = resistance
        target_by_2r = price + 2 * risk

        # For long: nearest upper target is TP1, farther upper target is TP2.
        tp1 = min(target_by_level, target_by_2r)
        tp2 = max(target_by_level, target_by_2r)

        rr1 = (tp1 - price) / risk
        rr2 = (tp2 - price) / risk

    elif side == "short":
        stop = max(resistance * 1.005, price + 1.2 * last["atr"])
        risk = stop - price

        if risk <= 0:
            return None

        target_by_level = support
        target_by_2r = price - 2 * risk

        # For short: nearest lower target is TP1, farther lower target is TP2.
        tp1 = max(target_by_level, target_by_2r)
        tp2 = min(target_by_level, target_by_2r)

        rr1 = (price - tp1) / risk
        rr2 = (price - tp2) / risk

    else:
        return None

    distance_to_support = (price - support) / price * 100
    distance_to_resistance = (resistance - price) / price * 100

    score = 0
    reasons = []

    if side == "long":
        if dlast["close"] > dlast["ema20"] > dlast["ema50"]:
            score += 2
            reasons.append("1D uptrend")

        if last["close"] > last["ema50"]:
            score += 1
            reasons.append("4H above EMA50")

        if 1 <= distance_to_support <= 8:
            score += 2
            reasons.append("near support")

        if 40 <= last["rsi"] <= 62:
            score += 1
            reasons.append("RSI not overheated")

        if distance_to_resistance < 2:
            score -= 2
            reasons.append("too close to resistance")

        if change24h > 8:
            score -= 2
            reasons.append("possible chase")

        btc_score_delta = btc_context.get("long_score_delta", 0)

        if btc_score_delta == 1:
            score += 1
            reasons.append("BTC supportive for long")
        elif btc_score_delta == -1:
            score -= 1
            reasons.append("BTC weak/neutral against long")
        elif btc_score_delta <= -2:
            score += btc_score_delta
            reasons.append("BTC bearish against long")

    elif side == "short":
        if dlast["close"] < dlast["ema20"] < dlast["ema50"]:
            score += 2
            reasons.append("1D downtrend")

        if last["close"] < last["ema50"]:
            score += 1
            reasons.append("4H below EMA50")

        if 1 <= distance_to_resistance <= 8:
            score += 2
            reasons.append("near resistance")

        if 38 <= last["rsi"] <= 65:
            score += 1
            reasons.append("RSI not oversold")

        if distance_to_support < 2:
            score -= 2
            reasons.append("too close to support")

        if change24h < -8:
            score -= 2
            reasons.append("possible late short")

        btc_score_delta = btc_context.get("short_score_delta", 0)

        if btc_score_delta > 0:
            score += btc_score_delta
            reasons.append("BTC supportive for short")
        elif btc_score_delta < 0:
            score += btc_score_delta
            reasons.append("BTC supportive against short")

    if rr1 >= 1.5:
        score += 2
        reasons.append("RR1 >= 1.5")

    if rr2 >= 2:
        score += 1
        reasons.append("RR2 >= 2")

    if side == "long" and funding < 0.04:
        score += 1
        reasons.append("funding ok")

    if side == "short" and funding > -0.04:
        score += 1
        reasons.append("funding ok")

    oi_context = analyze_open_interest(symbol, df4h, side)
    oi_score_delta = oi_context["oi_score_delta"]

    score += oi_score_delta

    if oi_score_delta > 0:
        reasons.append("OI confirms")
    elif oi_score_delta < 0:
        reasons.append("OI pressure")

    smc_context = analyze_smc(df4h, side)
    smc_score_delta = smc_context["smc_score_delta"]
    score += smc_score_delta

    if smc_context["smc_reason"]:
        reasons.append(smc_context["smc_reason"])

    return {
        "symbol": symbol,
        "side": side,
        "price": round(price, 8),
        "score": score,
        "entry": (
            f"{round(price * 0.995, 8)} - {round(price * 1.002, 8)}"
            if side == "long"
            else f"{round(price * 0.998, 8)} - {round(price * 1.005, 8)}"
        ),
        "stop": round(stop, 8),
        "tp1": round(tp1, 8),
        "tp2": round(tp2, 8),
        "rr1": round(rr1, 2),
        "rr2": round(rr2, 2),
        "turnover24h_m": round(turnover24h / 1_000_000, 1),
        "change24h_pct": round(change24h, 2),
        "funding_pct": round(funding, 4),
        "btc_regime": btc_context.get("regime", ""),
        "btc_score_delta": btc_score_delta,
        "oi_change_pct": oi_context["oi_change_pct"],
        "price_change_oi_window_pct": oi_context["price_change_oi_window_pct"],
        "oi_score_delta": oi_score_delta,
        "oi_signal": oi_context["oi_signal"],
        "smc_score_delta": smc_score_delta,
        "smc_bias": smc_context["smc_bias"],
        "smc_event": smc_context["smc_event"],
        "smc_zone": smc_context["smc_zone"],
        "smc_reason": smc_context["smc_reason"],
        "reason": ", ".join(reasons),
    }


def build_candidates(tickers):
    candidates = []

    for symbol, ticker in tickers.items():
        try:
            if not symbol.endswith("USDT"):
                continue

            turnover = float(ticker.get("turnover24h", 0))

            if turnover < MIN_TURNOVER:
                continue

            candidates.append({
                "symbol": symbol,
                "turnover": turnover,
            })

        except Exception:
            continue

    candidates = sorted(
        candidates,
        key=lambda x: x["turnover"],
        reverse=True,
    )

    return candidates[:MAX_SYMBOLS_TO_SCAN]


def append_csv_dedup(path, new_df, subset):
    if new_df.empty:
        return pd.DataFrame()

    if os.path.exists(path):
        old_df = pd.read_csv(path)
        final_df = pd.concat([old_df, new_df], ignore_index=True)
    else:
        final_df = new_df.copy()

    final_df = final_df.drop_duplicates(subset=subset, keep="last")
    final_df.to_csv(path, index=False, encoding="utf-8-sig")
    return final_df


def archive_existing_selected_results():
    if not os.path.exists(XLSX_FILE):
        return

    try:
        df = pd.read_excel(XLSX_FILE, engine="openpyxl")
    except Exception as e:
        print(f"Cannot archive previous selected signals: {e}")
        return

    if df.empty or "selected" not in df.columns or "signal_id" not in df.columns:
        return

    selected_df = df[df["selected"].apply(is_yes)].copy()

    if selected_df.empty:
        return

    selected_df["selected_saved_at"] = datetime.now(timezone.utc).isoformat()
    append_csv_dedup(SELECTED_SIGNALS_LOG, selected_df, ["signal_id"])

    if os.path.exists(SIGNALS_LOG):
        signals_df = pd.read_csv(SIGNALS_LOG)
        updates = selected_df.set_index("signal_id")
        signals_df = signals_df.copy()

        for col in ["selected", "manual_comment"]:
            if col not in signals_df.columns:
                signals_df[col] = ""
            if col in updates.columns:
                signals_df[col] = signals_df.apply(
                    lambda row: updates.loc[row["signal_id"], col]
                    if row["signal_id"] in updates.index
                    else row.get(col, ""),
                    axis=1,
                )

        signals_df.to_csv(SIGNALS_LOG, index=False, encoding="utf-8-sig")

    print(f"Archived selected signals to {SELECTED_SIGNALS_LOG}: {len(selected_df)}")


def append_signals_log(df):
    if df.empty:
        return

    append_csv_dedup(SIGNALS_LOG, df.copy(), ["signal_id"])
    print(f"Signals saved to {SIGNALS_LOG}")


def save_results(df):
    df.to_csv(
        CSV_FILE,
        index=False,
        encoding="utf-8-sig",
        sep=";",
        decimal=",",
    )

    with pd.ExcelWriter(XLSX_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Scanner")

        ws = writer.book["Scanner"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid",
        )

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

        header_map = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            header_map[cell.value] = col_idx

        text_columns = {
            "signal_id",
            "selected",
            "manual_comment",
            "created_at",
            "symbol",
            "side",
            "entry",
            "btc_regime",
            "btc_score_delta",
            "oi_signal",
            "smc_bias",
            "smc_event",
            "smc_zone",
            "smc_reason",
            "reason",
        }

        integer_columns = {
            "score",
            "oi_score_delta",
            "smc_score_delta",
        }

        decimal_columns = {
            "price",
            "stop",
            "tp1",
            "tp2",
            "rr1",
            "rr2",
            "turnover24h_m",
            "change24h_pct",
            "funding_pct",
            "oi_change_pct",
            "price_change_oi_window_pct",
        }

        for header, col_idx in header_map.items():
            col_letter = get_column_letter(col_idx)

            if header == "reason":
                ws.column_dimensions[col_letter].width = 80
            elif header == "manual_comment":
                ws.column_dimensions[col_letter].width = 45
            elif header in {"signal_id", "created_at", "oi_signal", "btc_regime", "smc_event"}:
                ws.column_dimensions[col_letter].width = 32
            elif header == "smc_reason":
                ws.column_dimensions[col_letter].width = 48
            elif header == "entry":
                ws.column_dimensions[col_letter].width = 26
            elif header == "selected":
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 16

            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue

                if header in text_columns:
                    cell.number_format = "@"
                    cell.alignment = Alignment(horizontal="left")
                elif header in integer_columns:
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="right")
                elif header in decimal_columns:
                    cell.number_format = "0.00000000"
                    cell.alignment = Alignment(horizontal="right")

    print("")
    print(f"Saved to {CSV_FILE}")
    print(f"Saved to {XLSX_FILE}")
    print("")
    print("Открывай именно XLSX, а не CSV.")
    print("В колонке selected поставь yes напротив монет, по которым выставил лимитки.")


def main():
    archive_existing_selected_results()

    btc_context = get_btc_context()
    market_context = get_market_regime_context()
    append_market_regime_log(market_context)

    print("")
    print("BTC context:")
    print(f"  regime: {btc_context['regime']}")
    print(f"  price: {btc_context['btc_price']}")
    print(f"  ema20 4H: {btc_context['btc_ema20']}")
    print(f"  ema50 4H: {btc_context['btc_ema50']}")
    print(f"  previous swing low: {btc_context['previous_swing_low']}")
    print("")

    print_market_regime_context(market_context)

    tickers = get_ticker_map()
    candidates = build_candidates(tickers)

    print(f"Scanning {len(candidates)} symbols...")
    print(f"Long signals enabled: {ENABLE_LONG_SIGNALS and btc_context['allow_longs']}")
    print(f"Short signals enabled: {ENABLE_SHORT_SIGNALS and btc_context['allow_shorts']}")

    results = []

    for item in candidates:
        symbol = item["symbol"]

        try:
            print(f"Scan {symbol}...")

            if ENABLE_LONG_SIGNALS and btc_context["allow_longs"]:
                long_result = analyze_symbol(symbol, tickers[symbol], btc_context, market_context, "long")
                if long_result:
                    results.append(long_result)

            if ENABLE_SHORT_SIGNALS and btc_context["allow_shorts"]:
                short_result = analyze_symbol(symbol, tickers[symbol], btc_context, market_context, "short")
                if short_result:
                    results.append(short_result)

            time.sleep(REQUEST_SLEEP)

        except Exception as e:
            print(f"Skip {symbol}: {e}")
            time.sleep(1)

    long_results = sorted(
        [row for row in results if row.get("side") == "long"],
        key=lambda x: x["score"],
        reverse=True,
    )[:TOP_N_PER_SIDE]

    short_results = sorted(
        [row for row in results if row.get("side") == "short"],
        key=lambda x: x["score"],
        reverse=True,
    )[:TOP_N_PER_SIDE]

    top = sorted(
        long_results + short_results,
        key=lambda x: (x["side"], -x["score"]),
    )

    df = pd.DataFrame(top)

    if df.empty:
        print("No setups found.")
        return

    created_at = datetime.now(timezone.utc).isoformat()
    created_compact = datetime.now().strftime("%Y%m%d_%H%M%S")

    df.insert(
        0,
        "signal_id",
        df.apply(
            lambda row: f"{created_compact}_{row['symbol']}_{str(row['side']).upper()}",
            axis=1,
        ),
    )

    df.insert(1, "selected", "")
    df.insert(2, "manual_comment", "")
    df.insert(3, "created_at", created_at)

    print("")
    print(tabulate(df, headers="keys", tablefmt="github", showindex=False))

    append_signals_log(df)
    save_results(df)


if __name__ == "__main__":
    main()
